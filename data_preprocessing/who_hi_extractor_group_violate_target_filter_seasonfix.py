"""
WHO HI assay Excel extractor

Converts WHO HI assay workbooks into a long-format table.
One workbook usually corresponds to an epidemic season, and one worksheet to one assay date.

Main output columns:
    date, virus, reference, reference_from_header, reference_from_row,
    virus_passage, virus_passage_type, reference_passage, reference_passage_type,
    virus_collection_date, HI_titre, censor, virus_seq, reference_seq

Extra QC columns:
    source_file, sheet_name, season, HI_raw, row, col,
    reference_match_score, date_is_season_fallback,
    virus_collection_date_is_inferred_from_name

Additional scored output:
    H1N1_WHO_HI_long_format_with_score_filtered.csv
    - removes records lacking virus_seq/reference_seq
    - removes records with >=4 X residues in either sequence
    - calculates a fixed-scale HI score using strict same-date self titres
    - keeps raw passage histories as virus_passage_raw/reference_passage_raw
"""

from __future__ import annotations

from pathlib import Path
import argparse
import glob
import random
import re
import sys
from datetime import datetime, date, timedelta
from difflib import SequenceMatcher
from typing import Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


# -----------------------------
# Basic string utilities
# -----------------------------

def cell_to_str(x) -> str:
    """Convert an Excel cell value to a normalized string for labels/names."""
    if x is None:
        return ""
    if isinstance(x, str):
        return re.sub(r"\s+", " ", x).strip()
    if isinstance(x, (datetime, date)):
        return x.strftime("%Y-%m-%d")
    return str(x).strip()


def norm_label(x) -> str:
    return re.sub(r"\s+", " ", cell_to_str(x).lower()).strip()


GLOBAL_HEADER_PATTERNS = [
    "haemagglutination inhibition",
    "hemagglutination inhibition",
    "post-infection",
    "post infection",
    "ferret antisera",
    "ferret antiserum",
    "ferret sera",
    "ferret serum",
]

REF_START_RE = re.compile(r"^(A|B)/|^IVR|^NIB|^NYMC|^X-|^RG-", re.I)

PASSAGE_RE = re.compile(
    r"(MDCK|SIAT|EGG|\bE\s*\d+\b|\bEX\b|CELL|AMNIOTIC|ALLANTOIC|C\d[,/]?E\d|P\d[/]?MDCK)",
    re.I,
)


def is_global_header(x) -> bool:
    s = norm_label(x)
    return any(p in s for p in GLOBAL_HEADER_PATTERNS)


# -----------------------------
# Virus-name normalization
# -----------------------------

ABBREV = {
    "/CAL/": "/CALIFORNIA/",
    "/CHCH/": "/CHRISTCHURCH/",
    "/HK/": "/HONGKONG/",
    "/ASTRAK/": "/ASTRAKHAN/",
    "/STP/": "/STPETERSBURG/",
    "/STHAFR/": "/SOUTHAFRICA/",
    "/NJ/": "/NEWJERSEY/",
    "/NJ2/": "/NEWJERSEY/",
    "/ENG/": "/ENGLAND/",
    "/AUCK/": "/AUCKLAND/",
    "/VIC/": "/VICTORIA/",
}


def expand_two_digit_year(yy: int) -> int:
    """
    Expand 2-digit years assuming all influenza collection years are between 1930 and 2029.

    Examples:
        76 -> 1976
        09 -> 2009
    """
    return 1900 + yy if 30 <= yy <= 99 else 2000 + yy


def strip_post_year_suffix_from_virus_name(name: str | None) -> Optional[str]:
    """
    Remove annotations appended after the strain-name year without mistaking
    year-like internal strain-name fields for the collection year.

    Rule:
        Slash-form strain names are interpreted as A/location/.../isolate/year.
        The year field is the rightmost slash-separated token that starts with
        a valid influenza year. This supports locations that themselves contain
        slashes, such as A/LYON/CHU/2182/1999, while avoiding truncation at
        year-like internal fields such as A/Panama/2007/1999.

        Anything after that year field is removed.

        Underscore-form names are handled analogously by taking the rightmost
        token that starts with a valid influenza year as the year field. This
        allows locations with underscores, e.g. A_HONG_KONG_2671_2019.

    Examples:
        A/LYON/CHU/2182/1999           -> A/LYON/CHU/2182/1999
        A/LYON/CHU/2182/1999@AA1A      -> A/LYON/CHU/2182/1999
        A/Panama/2007/1999             -> A/Panama/2007/1999
        A/Panama/2007/1999@AA1A        -> A/Panama/2007/1999
        A/Wisconsin/67/2005ISOLATE2    -> A/Wisconsin/67/2005
        A/Wisconsin/67/2005@AA1A       -> A/Wisconsin/67/2005
        A/Wisconsin/67/2005/1          -> A/Wisconsin/67/2005
        A/Trieste/25c/2007             -> A/Trieste/25c/2007
        A/St. Petersburg/1/2009        -> A/St Petersburg/1/2009
        A/O'Hare/1/2009                -> A/OHare/1/2009
        IVR-238 (A/Victoria/4897/2022@AA1A)
                                           -> IVR-238 (A/Victoria/4897/2022)
    """
    x = cell_to_str(name)
    if not x:
        return None

    x = x.replace("（", "(").replace("）", ")").strip()
    x = x.replace(".", "").replace("'", "")

    year_prefix_re = re.compile(r"^(19[3-9]\d|20[0-2]\d|\d{2})(.*)$")

    def clean_slash_strain(s: str) -> str:
        parts = s.split("/")
        if len(parts) < 4 or parts[0].upper() not in {"A", "B"}:
            return s

        # Use the rightmost slash-separated token that starts with a valid year.
        # This supports names with additional location fields, e.g.
        # A/LYON/CHU/2182/1999, while avoiding the internal year-like isolate in
        # A/Panama/2007/1999.
        best_i = None
        best_year = None
        for i in range(len(parts) - 1, 0, -1):
            m = year_prefix_re.match(parts[i].rstrip(")"))
            if m:
                best_i = i
                best_year = m.group(1)
                break

        if best_i is None:
            return s

        return "/".join(parts[:best_i] + [best_year])

    def clean_underscore_strain(s: str) -> str:
        parts = [p for p in s.split("_") if p != ""]
        if len(parts) < 4 or parts[0].upper() not in {"A", "B"}:
            return s

        # Underscore-form names may contain underscores in the location
        # component. Therefore, use the rightmost token that starts with a valid
        # year as the year field, then remove everything after it.
        best_i = None
        best_year = None
        for i in range(len(parts) - 1, 0, -1):
            m = year_prefix_re.match(parts[i].rstrip(")"))
            if m:
                best_i = i
                best_year = m.group(1)
                break

        if best_i is None:
            return s

        return "_".join(parts[:best_i] + [best_year])

    def clean_plain_strain(s: str) -> str:
        s = s.strip()
        if re.match(r"^[AB]/", s, flags=re.I):
            return clean_slash_strain(s)
        if re.match(r"^[AB]_", s, flags=re.I):
            return clean_underscore_strain(s)
        return s

    # Clean a strain name inside reassortant/vaccine notation first.
    # Example: IVR-238 (A/Victoria/4897/2022@AA1A)
    def replace_parenthetical(m: re.Match) -> str:
        return f"({clean_plain_strain(m.group(1))})"

    x = re.sub(r"\((A[/_][^)]+)\)", replace_parenthetical, x, flags=re.I)

    # If an annotation follows a cleaned parenthetical strain name, remove it.
    # Example: IVR-238 (A/Victoria/4897/2022)@note
    x = re.sub(r"^(.+\(A[/_][^)]+\)).+$", r"\1", x, flags=re.I)

    return clean_plain_strain(x)

def expand_years_in_name(name: str) -> str:
    """
    Expand only the rightmost slash/underscore-separated year token.

    Important: 2-digit-to-4-digit year expansion is applied only to the final
    strain-name year field, not to internal year-like fields.

    Examples:
        A/Memphis/102/72        -> A/Memphis/102/1972
        A/LYON/CHU/20/2000     -> A/LYON/CHU/20/2000
        A/LYON/CHU/20/00       -> A/LYON/CHU/20/2000
        IVR-238 (A/Vic/4897/22)-> IVR-238 (A/Vic/4897/2022)
    """
    x = strip_post_year_suffix_from_virus_name(name)

    if not x:
        return ""

    def expand_last_token(s: str) -> str:
        s = s.strip()
        close = ")" if s.endswith(")") else ""
        core = s[:-1] if close else s

        sep = "/" if re.match(r"^[AB]/", core, flags=re.I) else "_" if re.match(r"^[AB]_", core, flags=re.I) else None
        if sep is None:
            # Non-standard fallback: only expand a truly terminal /YY token.
            m = re.search(r"/(\d{2})$", core)
            if m:
                return core[:m.start()] + f"/{expand_two_digit_year(int(m.group(1)))}" + close
            return s

        parts = core.split(sep) if sep == "/" else [p for p in core.split("_") if p != ""]
        if len(parts) >= 4 and re.fullmatch(r"\d{2}", parts[-1]):
            parts[-1] = f"{expand_two_digit_year(int(parts[-1]))}"
            core = sep.join(parts)
        return core + close

    # First expand strain names inside reassortant/vaccine parentheses.
    x = re.sub(
        r"\((A[/_][^)]+)\)",
        lambda m: f"({expand_last_token(m.group(1))})",
        x,
        flags=re.I,
    )

    # Then expand if the whole value itself is a strain name.
    return expand_last_token(x)


def standardize_virus_name(name: str | None) -> Optional[str]:
    """
    Standardize virus strain names for output and FASTA matching.

    - Expand 2-digit years to 1930-2029.
    - Replace spaces with underscores.
    """
    x = cell_to_str(name)

    if not x:
        return None

    x = x.replace("（", "(").replace("）", ")")
    x = expand_years_in_name(x)
    x = re.sub(r"\s+", "_", x.strip())

    return x


def canonical_virus_name(name: str | None) -> str:
    """
    Canonicalize virus names for fuzzy matching of abbreviated reference headers.

    Any suffix appended after the terminal strain-name year is removed first.
    """
    x = strip_post_year_suffix_from_virus_name(name)

    if not x:
        return ""

    x = x.upper()
    x = x.replace("（", "(").replace("）", ")")
    x = re.sub(r"\s+", "", x)
    x = x.replace("_", "")
    x = x.replace("-", "")

    for k, v in ABBREV.items():
        x = x.replace(k.replace(" ", ""), v.replace(" ", ""))

    def repl_year(m):
        yy = int(m.group(1))
        return f"/{expand_two_digit_year(yy)}"

    x = re.sub(r"/(\d{2})(?=\s*(?:$|\)))", repl_year, x)
    x = re.sub(r"[^A-Z0-9/()]", "", x)

    return x


def candidate_match_names(name: str | None) -> list[str]:
    """
    Names/aliases to try for sequence lookup.
    """
    out = []

    name_clean = strip_post_year_suffix_from_virus_name(name)

    if name_clean:
        out.append(name_clean)

        # Reassortant / vaccine-strain notation:
        # IVR-238_(A/Victoria/4897/2022) -> also try A/Victoria/4897/2022
        m = re.search(r"\((A/[^)]+)\)", cell_to_str(name_clean), flags=re.I)
        if m:
            out.append(m.group(1))

    seen = set()
    candidates = []

    for x in out:
        sx = standardize_virus_name(x)

        if sx and sx not in seen:
            seen.add(sx)
            candidates.append(sx)

    return candidates


# -----------------------------
# Header detection
# -----------------------------

def find_virus_col(ws, max_rows=30):
    """
    Find the row/column of the 'Viruses' header.
    """
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            s = norm_label(ws.cell(r, c).value)

            if s in {"virus", "viruses"}:
                return r, c

    raise ValueError(f"Cannot find virus column in sheet: {ws.title}")


def find_collection_col(ws, max_rows=30):
    """
    Find Collection Date column, allowing 'Collection' and 'Date' split across two rows.
    """
    # Case 1: 'Collection Date' in one cell.
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            s = norm_label(ws.cell(r, c).value)

            if "collection date" in s or "collection/date" in s:
                return r, c

    # Case 2: 'Collection' in one row and 'Date' below it.
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            s = norm_label(ws.cell(r, c).value)

            if s == "collection" or s.startswith("collection "):
                for rr in range(r, min(r + 3, ws.max_row) + 1):
                    t = norm_label(ws.cell(rr, c).value)

                    if t == "date" or t.startswith("date"):
                        return r, c

    raise ValueError(f"Cannot find Collection Date column in sheet: {ws.title}")


def find_section_row(ws, virus_col, label, start_row=1):
    """
    Find section labels such as 'REFERENCE VIRUSES' or 'TEST VIRUSES'.
    """
    label_norm = label.lower()

    for r in range(start_row, ws.max_row + 1):
        s = norm_label(ws.cell(r, virus_col).value)

        if label_norm in s:
            return r

    return None


def find_reference_name_row(ws, passage_col, ref_section_row=None, max_header_rows=25):
    """
    Find the first row of the two-row reference-virus header.
    """
    end_row = min((ref_section_row or max_header_rows) - 1, max_header_rows, ws.max_row)
    best = None

    for r in range(1, end_row + 1):
        count_nonempty = 0
        count_ref_like = 0

        for c in range(passage_col + 1, ws.max_column + 1):
            s = cell_to_str(ws.cell(r, c).value)

            if not s or is_global_header(s):
                continue

            count_nonempty += 1

            if REF_START_RE.search(s):
                count_ref_like += 1

        if count_ref_like > 0:
            score = (count_ref_like, count_nonempty, -r)

            if best is None or score > best[0]:
                best = (score, r)

    if best is None:
        raise ValueError(f"Cannot find reference-virus header row in sheet: {ws.title}")

    return best[1]


def combine_reference_name(part1, part2):
    """
    Combine the two-row reference name into one string.
    """
    a = cell_to_str(part1)
    b = cell_to_str(part2)

    if not a and not b:
        return None

    if not a:
        return b

    if not b:
        return a

    # A/Cal + 7/09 -> A/Cal/7/09
    # A/Victoria/ + 4897/2022 -> A/Victoria/4897/2022
    if a.endswith("/") or b.startswith("/"):
        return f"{a}{b}"

    if re.match(r"^[AB]/", a, flags=re.I) and re.match(r"^[A-Za-z0-9-]+/", b):
        return f"{a}/{b}"

    if re.match(r"^[AB]/", a, flags=re.I) and re.match(r"^\d", b):
        return f"{a}/{b}"

    # Reassortants / vaccine strains:
    # IVR-238 + A/Vic/4897/22 -> IVR-238 (A/Vic/4897/22)
    return f"{a} ({b})"


def looks_like_passage(s):
    return bool(PASSAGE_RE.search(cell_to_str(s)))


def classify_passage_type(passage):
    """
    Classify passage information as EGG, CELL, EGG|CELL, CELL|EGG, or UNKNOWN.

    Rules:
        Empty / NA-like values -> UNKNOWN
        Egg, E1, E2, E3, E4, Ex, Am/Al, etc. -> EGG
        SIAT, MDCK, C1, P1, etc. -> CELL
        Mixed histories preserve the observed order:
            E3/SIAT  -> EGG|CELL
            SIAT/E3  -> CELL|EGG
            C1/E2    -> CELL|EGG
            E2/C1    -> EGG|CELL
    """
    s = cell_to_str(passage)

    if not s:
        return "UNKNOWN"

    s_norm = s.strip().upper()

    if s_norm in {"NA", "N/A", "N.D.", "ND", "UNKNOWN", "UNK", "NONE", "-", "--", "?"}:
        return "UNKNOWN"

    s_norm = s_norm.replace(" ", "")

    egg_patterns = [
        r"EGG",
        r"EMBRYONATEDEGG",
        r"E\d+",
        r"EX",
        r"AMNIOTIC",
        r"ALLANTOIC",
        r"AM\d+",
        r"AL\d+",
    ]
    cell_patterns = [
        r"CELL",
        r"MDCK",
        r"SIAT",
        r"C\d+",
        r"P\d+",
    ]

    def first_match_pos(token: str, patterns: list[str]) -> Optional[int]:
        positions = []
        for pat in patterns:
            m = re.search(pat, token)
            if m:
                positions.append(m.start())
        return min(positions) if positions else None

    def token_to_types(token: str) -> list[str]:
        if not token:
            return []

        egg_pos = first_match_pos(token, egg_patterns)
        cell_pos = first_match_pos(token, cell_patterns)

        if egg_pos is not None and cell_pos is not None:
            if egg_pos < cell_pos:
                return ["EGG", "CELL"]
            if cell_pos < egg_pos:
                return ["CELL", "EGG"]
            # Extremely rare ambiguous case where both patterns start at the same position.
            # Prefer EGG because E/Ex/egg terms are more specific than generic passage counts.
            return ["EGG"]

        if egg_pos is not None:
            return ["EGG"]

        if cell_pos is not None:
            return ["CELL"]

        # Preserve previous extractor behavior: any other non-empty passage label
        # is treated as CELL rather than UNKNOWN.
        return ["CELL"]

    normalized = []
    for part in re.split(r"[|;/,+]+", s_norm):
        for passage_type in token_to_types(part):
            if passage_type not in normalized:
                normalized.append(passage_type)

    if not normalized:
        return "UNKNOWN"

    return "|".join(normalized)

def detect_reference_passage_header_row(ws, name_row, ref_cols):
    """
    Detect a reference-passage row directly under the reference names.
    """
    for r in [name_row + 2, name_row + 3]:
        vals = [cell_to_str(ws.cell(r, c).value) for c in ref_cols]
        vals = [v for v in vals if v]

        if not vals:
            continue

        ratio = sum(looks_like_passage(v) for v in vals) / len(vals)

        if ratio >= 0.5:
            return r

    return None


def get_reference_columns(ws, passage_col, name_row, data_start_row=None):
    """
    Return columns that correspond to reference antisera / HI titre columns.
    """
    ref_cols = []

    for c in range(passage_col + 1, ws.max_column + 1):
        ref_name = combine_reference_name(ws.cell(name_row, c).value, ws.cell(name_row + 1, c).value)

        if not ref_name:
            continue

        if not REF_START_RE.search(ref_name) and "/" not in ref_name:
            continue

        # Keep only columns with at least one HI-like value below the header.
        if data_start_row is not None:
            has_data = False

            for r in range(data_start_row, min(ws.max_row, data_start_row + 100) + 1):
                titer, censor, raw = parse_hi_value(ws.cell(r, c).value)

                if titer is not None or censor is not None:
                    has_data = True
                    break

            if not has_data:
                continue

        ref_cols.append(c)

    return ref_cols


# -----------------------------
# Reference-virus block mapping
# -----------------------------

def build_reference_passage_rows(ws, virus_col, passage_col, ref_section_row):
    """
    Collect full reference-virus names and passage histories from the REFERENCE VIRUSES block.
    """
    rows = []

    if ref_section_row is None:
        return rows

    blank_run = 0

    for r in range(ref_section_row + 1, ws.max_row + 1):
        virus = cell_to_str(ws.cell(r, virus_col).value)

        if "TEST VIRUSES" in virus.upper():
            break

        if not virus:
            blank_run += 1

            if blank_run >= 3 and rows:
                break

            continue

        blank_run = 0

        if "VIRUSES" in virus.upper() or "COPY AND PASTE" in virus.upper():
            continue

        rows.append(
            {
                "name": virus,
                "standard_name": standardize_virus_name(virus),
                "passage": cell_to_str(ws.cell(r, passage_col).value) or None,
                "row": r,
                "canon": canonical_virus_name(virus),
            }
        )

    return rows


def best_reference_match(ref_name, ref_rows, min_score=0.55):
    """
    Fuzzy-match a reference header name to the full REFERENCE VIRUSES row.
    """
    if not ref_rows or not ref_name:
        return None, None

    queries = [canonical_virus_name(ref_name)]

    # If header is like IVR-238 (A/Vic/4897/22), also try the parent virus inside parentheses.
    m = re.search(r"\((A/[^)]+)\)", ref_name, flags=re.I)

    if m:
        queries.append(canonical_virus_name(m.group(1)))

    best_score = 0.0
    best_row = None

    for q in queries:
        for row in ref_rows:
            score = SequenceMatcher(None, q, row["canon"]).ratio()

            if q in row["canon"] or row["canon"] in q:
                score = max(score, 0.90)

            if score > best_score:
                best_score = score
                best_row = row

    if best_row is not None and best_score >= min_score:
        return best_row, best_score

    return None, best_score if best_score else None


# -----------------------------
# Value parsers
# -----------------------------

def parse_hi_value(value):
    """
    Parse HI titre values.

    Censor rule:
        640      -> HI_titre=640,  censor=0
        '<40'    -> HI_titre=40,   censor=1
        '<'      -> HI_titre=40,   censor=1
        '>5120'  -> HI_titre=5120, censor=0
        '>=5120' -> HI_titre=5120, censor=0
    """
    raw = cell_to_str(value)

    if raw == "":
        return None, None, None

    s = raw.replace(",", "").replace("≤", "<=").replace("≥", ">=").strip()

    # Only lower-direction inequalities are treated as censored.
    # Upper-direction inequalities are ignored for the censor flag.
    censor = 1 if "<" in s else 0

    # Rare old-format cells may contain only '<'. Read them as '<40'.
    if s in {"<", "<="}:
        return 40, 1, raw

    m = re.search(r"(\d+(?:\.\d+)?)", s)

    if not m:
        return None, censor if censor else None, raw

    num = float(m.group(1))

    if num.is_integer():
        num = int(num)

    return num, censor, raw


def parse_month_year_string(s: str) -> Optional[str]:
    """
    Parse month-year strings such as Nov-09, Nov/09, Nov 09, November-2009.

    These are converted to the first day of the month:
        Nov-09 -> 2009-11-01

    2-digit years are expanded using the 1930-2029 rule:
        30-99 -> 1930-1999
        00-29 -> 2000-2029
    """
    s = cell_to_str(s)

    if not s:
        return None

    month_map = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }

    # Month-year, e.g. Nov-09, Nov/09, Nov 09, November-2009
    m = re.fullmatch(
        r"\s*([A-Za-z]{3,9})[\s\-/\.]+(\d{2}|\d{4})\s*",
        s,
        flags=re.I,
    )

    if m:
        mon_s = m.group(1).lower()
        yy_s = m.group(2)

        if mon_s in month_map:
            month = month_map[mon_s]
            year = int(yy_s)

            if year < 100:
                year = expand_two_digit_year(year)

            if 1930 <= year <= 2029:
                return f"{year:04d}-{month:02d}-01"

    # Year-month with month name, e.g. 2009-Nov, 09-Nov
    m = re.fullmatch(
        r"\s*(\d{2}|\d{4})[\s\-/\.]+([A-Za-z]{3,9})\s*",
        s,
        flags=re.I,
    )

    if m:
        yy_s = m.group(1)
        mon_s = m.group(2).lower()

        if mon_s in month_map:
            month = month_map[mon_s]
            year = int(yy_s)

            if year < 100:
                year = expand_two_digit_year(year)

            if 1930 <= year <= 2029:
                return f"{year:04d}-{month:02d}-01"

    return None


def parse_date_string(s):
    """
    Parse common date strings without throwing warnings for DD/MM/YYYY-like dates.

    Month-year strings such as Nov-09 are interpreted as the first day of that month:
        Nov-09 -> 2009-11-01
    """
    s = cell_to_str(s)

    if not s:
        return None

    # First handle month-year strings such as Nov-09.
    month_year = parse_month_year_string(s)

    if month_year:
        return month_year

    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%b %d %Y",
        "%B %d %Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass

    # Fallback: avoid dayfirst=True warning.
    # DD/MM/YYYY is already handled explicitly above, so dayfirst=False is safer here.
    try:
        return pd.to_datetime(s, errors="raise", dayfirst=False).date().isoformat()
    except Exception:
        return None


def to_iso_date(value, workbook_epoch=None):
    if value is None or cell_to_str(value) == "":
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)) and 20000 <= value <= 60000 and workbook_epoch is not None:
        try:
            return from_excel(value, epoch=workbook_epoch).date().isoformat()
        except Exception:
            pass

    return parse_date_string(value)


def extract_year_from_virus_name(virus_name):
    """
    Extract collection year from a virus strain name.

    Examples:
        A/New_Jersey/8/1976             -> 1976
        A/New_Jersey/8/76               -> 1976
        A/California/7/09               -> 2009
        A/LYON/CHU/2182/1999            -> 1999
        A/Panama/2007/1999              -> 1999
        IVR-238 (A/LYON/CHU/20/00)      -> 2000

    Any suffix appended after the strain-name year is removed first.
    For slash-form names, only the rightmost slash-separated token is treated
    as the strain year; internal year-like tokens are ignored.
    """
    x = strip_post_year_suffix_from_virus_name(virus_name)

    if not x:
        return None

    x = x.replace("（", "(").replace("）", ")")

    def year_from_token(token: str) -> Optional[int]:
        token = token.rstrip(")")
        if re.fullmatch(r"19[3-9]\d|20[0-2]\d", token):
            return int(token)
        if re.fullmatch(r"\d{2}", token):
            return expand_two_digit_year(int(token))
        return None

    def year_from_delimited_strain(s: str) -> Optional[int]:
        s = s.strip()
        if re.match(r"^[AB]/", s, flags=re.I):
            parts = s.split("/")
            if len(parts) >= 4:
                return year_from_token(parts[-1])
        if re.match(r"^[AB]_", s, flags=re.I):
            parts = [p for p in s.split("_") if p != ""]
            if len(parts) >= 4:
                return year_from_token(parts[-1])
        return None

    # First handle reassortant/vaccine notation by inspecting only the strain
    # inside parentheses, e.g. IVR-238 (A/LYON/CHU/20/00).
    m = re.search(r"\((A[/_][^)]+)\)", x, flags=re.I)
    if m:
        year = year_from_delimited_strain(m.group(1))
        if year is not None:
            return year

    # Standard slash/underscore-form names: only the final token is the year.
    year = year_from_delimited_strain(x)
    if year is not None:
        return year

    # Fallback for non-standard names: use only a truly terminal 4-digit or
    # 2-digit year token, not an internal /20/ field.
    m4 = re.search(r"(?:^|[/_\s(])(19[3-9]\d|20[0-2]\d)(?:\)?$)", x)
    if m4:
        return int(m4.group(1))

    m2 = re.search(r"(?:^|[/_\s(])(\d{2})(?:\)?$)", x)
    if m2:
        return expand_two_digit_year(int(m2.group(1)))

    return None

def fill_collection_date_from_virus_name(collection_date, virus_name):
    """
    If virus_collection_date is missing, infer it from the year in the virus name.

    Example:
        collection_date = None
        virus_name = A/New_Jersey/8/1976
        -> 1976-01-01
    """
    if collection_date:
        return collection_date, False

    year = extract_year_from_virus_name(virus_name)

    if year is None:
        return None, False

    return f"{year:04d}-01-01", True


def extract_dates_from_text(text: str) -> list[str]:
    """
    Extract dates from free text and return ISO dates.

    Supports:
        YYYY-MM-DD
        DD/MM/YYYY
        DD Mon YYYY
        Nov-09 -> 2009-11-01
    """
    if not text:
        return []

    s = cell_to_str(text)
    candidates = []

    # ISO-like or slash/dash dates with 4-digit year first.
    candidates += re.findall(r"\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b", s)

    # DD/MM/YYYY or DD-MM-YYYY.
    candidates += re.findall(r"\b\d{1,2}[-/.]\d{1,2}[-/.]\d{4}\b", s)

    # DD Mon YYYY / Mon DD YYYY.
    mon = r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*"
    candidates += re.findall(rf"\b\d{{1,2}}\s+{mon}\s+\d{{4}}\b", s, flags=re.I)
    candidates += re.findall(rf"\b{mon}\s+\d{{1,2}},?\s+\d{{4}}\b", s, flags=re.I)

    # Month-year strings such as Nov-09, Nov/09, Nov 09, November-2009.
    candidates += re.findall(rf"\b{mon}[\s\-/\.]+(?:\d{{2}}|\d{{4}})\b", s, flags=re.I)

    # Year-month with month name, e.g. 2009-Nov.
    candidates += re.findall(rf"\b(?:\d{{2}}|\d{{4}})[\s\-/\.]+{mon}\b", s, flags=re.I)

    parsed = []

    for cand in candidates:
        iso = parse_date_string(cand.replace(",", ""))

        if iso:
            parsed.append(iso)

    return parsed


def extract_table_title_date(ws, max_rows=12, max_cols=20):
    """
    Extract the assay date from the table title.

    Primary rule:
        Use the top-left cell (A1). WHO tables usually store the assay-date
        title there, for example:
            Table 5-11. Antigenic analyses ... (2018-02-07)
            Table 5. Antigenic analyses ... (09/02/2010)

    Fallback:
        If A1 has no parseable date, scan the nearby header cells in row-major
        order and return the first date found in the first cell that contains a
        date. This avoids the previous behavior of collecting all dates from the
        header/data area and returning the earliest one, which could accidentally
        select an old virus collection date such as 2021-04-16.
    """

    def dates_from_cell(value) -> list[str]:
        if isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                return [value.date().isoformat()]
            return [value.isoformat()]
        return extract_dates_from_text(cell_to_str(value))

    # Prefer A1 exactly.
    a1_dates = dates_from_cell(ws.cell(1, 1).value)
    if a1_dates:
        return a1_dates[0]

    # Fallback: scan nearby cells, but stop at the first cell with a date.
    # Do not aggregate dates across cells or choose min(dates), because that can
    # capture collection dates from early data rows rather than the table title.
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, min(max_cols, ws.max_column) + 1):
            if r == 1 and c == 1:
                continue
            dates = dates_from_cell(ws.cell(r, c).value)
            if dates:
                return dates[0]

    return None


def season_from_path(path) -> Optional[str]:
    stem = Path(path).stem
    m = re.search(r"\b([SN]H\d{4})\b", stem, flags=re.I)
    return m.group(1).upper() if m else None


def derive_season_from_date(date_value: str | None) -> Optional[str]:
    """
    Derive influenza season using the updated naming convention:
        SHYYYY: YYYY-02-01 to YYYY-08-31
        NHYYYY: (YYYY-1)-09-01 to YYYY-01-31

    Examples:
        2014-09-01 -> NH2015
        2015-01-31 -> NH2015
        2015-02-01 -> SH2015
        2015-09-01 -> NH2016
    """
    if not date_value:
        return None

    try:
        d = pd.to_datetime(date_value, errors="raise").date()
    except Exception:
        return None

    if 2 <= d.month <= 8:
        return f"SH{d.year}"
    if 9 <= d.month <= 12:
        return f"NH{d.year + 1}"
    if d.month == 1:
        return f"NH{d.year}"
    return None


def season_fallback_start_date(season: str | None) -> Optional[date]:
    """
    Return the fallback start date for an epidemic season.

    Updated season definition:
        SHYYYY: YYYY-02-01 to YYYY-08-31
        NHYYYY: (YYYY-1)-09-01 to YYYY-01-31

    If table-title date is missing, fallback dates are assigned sequentially from:
        SHYYYY -> YYYY-02-01, YYYY-02-02, ...
        NHYYYY -> (YYYY-1)-09-01, (YYYY-1)-09-02, ...
    """
    if not season:
        return None

    m = re.match(r"^([SN]H)(\d{4})$", season.upper())

    if not m:
        return None

    hemi, year_s = m.groups()
    year = int(year_s)

    if hemi == "SH":
        return date(year, 2, 1)

    if hemi == "NH":
        return date(year - 1, 9, 1)

    return None


# -----------------------------
# FASTA handling
# -----------------------------

def looks_like_influenza_strain_name(x: str | None) -> bool:
    """
    Return True if a FASTA header field looks like an influenza strain name.

    Examples:
        A/Kamata/85/1987      -> True
        A/Memphis/102/72      -> True
        B/Victoria/2/1987     -> True
        A_/_H3N2              -> False
        EPI_ISL_21051         -> False
        HA                    -> False
    """
    s = cell_to_str(x)

    if not s:
        return False

    if not re.match(r"^[AB]/", s, flags=re.I):
        return False

    # Require a plausible collection year in the strain name.
    return extract_year_from_virus_name(s) is not None


def extract_fasta_strain_name(header: str | None) -> Optional[str]:
    """
    Extract strain name from FASTA header.

    Supports both header formats:

    H1N1-style:
        >EPI275|HA|A/Kamata/85/1987|EPI_ISL_107|A_/_H1N1
        -> A/Kamata/85/1987

    H3N2-style:
        >A/Memphis/102/72|EPI_ISL_21051|A_/_H3N2|||unassigned|1972-01-01|HA|EPI118986
        -> A/Memphis/102/72

    Strategy:
        1. Split header by "|".
        2. Search all fields for an A/... or B/... strain name with a plausible year.
        3. If not found, fall back to older H1N1-style parts[2].
        4. If still not found, fall back to parts[0].
    """
    if not header:
        return None

    parts = [cell_to_str(p) for p in header.split("|")]

    # Best strategy: search all fields for strain-like names.
    for p in parts:
        if looks_like_influenza_strain_name(p):
            return p

    # Fallback for old H1N1-style headers.
    if len(parts) >= 3 and parts[2]:
        return parts[2]

    # Fallback for H3N2-style headers or simple FASTA headers.
    if len(parts) >= 1 and parts[0]:
        return parts[0].split()[0]

    return None


def read_fasta_sequences(fasta_path: str | Path | None) -> dict[str, str]:
    """
    Read FASTA and map standardized strain name to sequence.

    Supported FASTA header examples:

        H1N1-style:
            >EPI275|HA|A/Kamata/85/1987|EPI_ISL_107|A_/_H1N1

        H3N2-style:
            >A/Memphis/102/72|EPI_ISL_21051|A_/_H3N2|||unassigned|1972-01-01|HA|EPI118986

    The strain name is auto-detected from the header.

    Gap characters '-' are removed from sequences.

    If multiple FASTA records map to the same standardized strain name, select
    a sequence with the fewest X residues. If multiple records tie for the
    lowest X count, randomly choose one of the tied records.
    """
    if fasta_path is None:
        return {}

    fasta_path = Path(fasta_path)
    records_by_key: dict[str, list[str]] = {}
    current_name = None
    current_seq = []

    def clean_sequence(seq_chunks: list[str]) -> str:
        return (
            "".join(seq_chunks)
            .replace(" ", "")
            .replace("\n", "")
            .replace("\r", "")
            .replace("-", "")
            .strip()
        )

    def flush():
        nonlocal current_name, current_seq

        if current_name and current_seq:
            key = standardize_virus_name(current_name)
            seq = clean_sequence(current_seq)

            if key and seq:
                records_by_key.setdefault(key, []).append(seq)

        current_name = None
        current_seq = []

    with fasta_path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")

            if line.startswith(">"):
                flush()
                header = line[1:].strip()
                current_name = extract_fasta_strain_name(header)
            else:
                current_seq.append(line.strip())

    flush()

    seqs: dict[str, str] = {}
    for key, seq_list in records_by_key.items():
        min_x = min(seq.upper().count("X") for seq in seq_list)
        best = [seq for seq in seq_list if seq.upper().count("X") == min_x]
        seqs[key] = random.choice(best)

    return seqs

def lookup_sequence(name: str | None, seq_map: dict[str, str]) -> Optional[str]:
    if not seq_map:
        return None

    for cand in candidate_match_names(name):
        if cand in seq_map:
            return seq_map[cand]

    return None


# -----------------------------
# Main extraction functions
# -----------------------------

def extract_sheet(ws, source_file):
    virus_header_row, virus_col = find_virus_col(ws)
    collection_row, collection_col = find_collection_col(ws)
    passage_col = collection_col + 1

    ref_section_row = find_section_row(ws, virus_col, "REFERENCE VIRUSES")
    name_row = find_reference_name_row(ws, passage_col, ref_section_row=ref_section_row)
    data_start_row = (ref_section_row + 1) if ref_section_row else max(virus_header_row, collection_row) + 1

    ref_cols = get_reference_columns(ws, passage_col, name_row, data_start_row=data_start_row)
    reference_passage_header_row = detect_reference_passage_header_row(ws, name_row, ref_cols)
    reference_passage_rows = build_reference_passage_rows(ws, virus_col, passage_col, ref_section_row)

    sheet_date = extract_table_title_date(ws)
    # Prefer the actual assay date for season assignment. Fall back to the
    # filename season only when no parseable table-title date is available.
    season = derive_season_from_date(sheet_date) or season_from_path(source_file)

    ref_meta = {}

    for c in ref_cols:
        reference_from_header_raw = combine_reference_name(
            ws.cell(name_row, c).value,
            ws.cell(name_row + 1, c).value,
        )

        reference_from_header = standardize_virus_name(reference_from_header_raw)

        match_row, match_score = best_reference_match(
            reference_from_header_raw,
            reference_passage_rows,
        )

        reference_from_row = match_row["standard_name"] if match_row else None

        reference_passage = None

        if reference_passage_header_row is not None:
            reference_passage = cell_to_str(ws.cell(reference_passage_header_row, c).value) or None

        if not reference_passage and match_row:
            reference_passage = match_row["passage"]

        ref_meta[c] = {
            "reference": reference_from_row or reference_from_header,
            "reference_from_header": reference_from_header,
            "reference_from_row": reference_from_row,
            "reference_match_score": match_score,
            "reference_passage": reference_passage,
        }

    extracted = []

    for r in range(data_start_row, ws.max_row + 1):
        virus_raw = cell_to_str(ws.cell(r, virus_col).value)

        if not virus_raw:
            continue

        virus_upper = virus_raw.upper()

        if "VIRUSES" in virus_upper or "COPY AND PASTE" in virus_upper:
            continue

        # Skip footnotes/legend rows that may appear below the data table.
        # Test-virus rows should be strain-like names, e.g. A/... or B/... .
        if not REF_START_RE.search(virus_raw) and "/" not in virus_raw:
            continue

        virus = standardize_virus_name(virus_raw)
        virus_passage = cell_to_str(ws.cell(r, passage_col).value) or None

        virus_collection_date = to_iso_date(
            ws.cell(r, collection_col).value,
            workbook_epoch=ws.parent.epoch,
        )

        # Important:
        # If virus_collection_date is missing, infer YYYY-01-01 from the virus name.
        virus_collection_date, virus_collection_date_is_inferred_from_name = fill_collection_date_from_virus_name(
            virus_collection_date,
            virus_raw,
        )

        for c in ref_cols:
            hi_titre, censor, raw_hi = parse_hi_value(ws.cell(r, c).value)

            if hi_titre is None and censor is None:
                continue

            extracted.append(
                {
                    "source_file": Path(source_file).name,
                    "sheet_name": ws.title,
                    "season": season,
                    "date": sheet_date,
                    "date_is_season_fallback": False,
                    "virus": virus,
                    "reference": ref_meta[c]["reference"],
                    "reference_from_header": ref_meta[c]["reference_from_header"],
                    "reference_from_row": ref_meta[c]["reference_from_row"],
                    "reference_match_score": ref_meta[c]["reference_match_score"],
                    "virus_passage": virus_passage,
                    "virus_passage_type": classify_passage_type(virus_passage),
                    "reference_passage": ref_meta[c]["reference_passage"],
                    "reference_passage_type": classify_passage_type(ref_meta[c]["reference_passage"]),
                    "virus_collection_date": virus_collection_date,
                    "virus_collection_date_is_inferred_from_name": virus_collection_date_is_inferred_from_name,
                    "HI_titre": hi_titre,
                    "censor": censor,
                    "HI_raw": raw_hi,
                    "row": r,
                    "col": c,
                }
            )

    sheet_log = {
        "source_file": Path(source_file).name,
        "sheet_name": ws.title,
        "season": season,
        "date": sheet_date,
        "date_is_season_fallback": False,
        "virus_col": virus_col,
        "collection_col": collection_col,
        "passage_col": passage_col,
        "reference_name_row": name_row,
        "reference_passage_header_row": reference_passage_header_row,
        "n_reference_cols": len(ref_cols),
        "n_extracted_rows": len(extracted),
    }

    return pd.DataFrame(extracted), sheet_log


def apply_season_date_fallback(df: pd.DataFrame, log_df: pd.DataFrame, source_file) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Assign sequential fallback dates for sheets without table-title dates.
    """
    season = season_from_path(source_file)
    start = season_fallback_start_date(season)

    # Avoid pandas FutureWarning:
    # assigning date strings to a column initially inferred as float64.
    if not df.empty and "date" in df.columns:
        df["date"] = df["date"].astype("object")

    if not log_df.empty and "date" in log_df.columns:
        log_df["date"] = log_df["date"].astype("object")

    if start is None or log_df.empty:
        return df, log_df

    if "date" in log_df.columns:
        date_series = log_df["date"]
    else:
        date_series = pd.Series([pd.NA] * len(log_df), index=log_df.index)

    if "error" in log_df.columns:
        error_series = log_df["error"]
    else:
        error_series = pd.Series([pd.NA] * len(log_df), index=log_df.index)

    missing_logs = log_df[date_series.isna() & error_series.isna()]

    fallback_by_sheet = {}

    for i, (_, row) in enumerate(missing_logs.iterrows()):
        fallback_by_sheet[row["sheet_name"]] = (start + timedelta(days=i)).isoformat()

    if fallback_by_sheet:
        for sheet_name, fallback_date in fallback_by_sheet.items():
            fallback_season = derive_season_from_date(fallback_date) or season

            if not df.empty:
                mask = (
                    (df["source_file"] == Path(source_file).name)
                    & (df["sheet_name"] == sheet_name)
                    & df["date"].isna()
                )

                df.loc[mask, "date"] = fallback_date
                df.loc[mask, "date_is_season_fallback"] = True
                if "season" in df.columns:
                    df.loc[mask, "season"] = fallback_season

            log_mask = (
                (log_df["source_file"] == Path(source_file).name)
                & (log_df["sheet_name"] == sheet_name)
            )

            log_df.loc[log_mask, "date"] = fallback_date
            log_df.loc[log_mask, "date_is_season_fallback"] = True
            if "season" in log_df.columns:
                log_df.loc[log_mask, "season"] = fallback_season

    # Recalculate season from every valid assay date. This keeps outputs correct
    # even when input workbook filenames still use the older NH naming convention.
    if not df.empty and "date" in df.columns:
        if "season" not in df.columns:
            df["season"] = pd.NA
        derived = df["date"].apply(derive_season_from_date)
        mask = derived.notna()
        df.loc[mask, "season"] = derived[mask]

    if not log_df.empty and "date" in log_df.columns:
        if "season" not in log_df.columns:
            log_df["season"] = pd.NA
        derived = log_df["date"].apply(derive_season_from_date)
        mask = derived.notna()
        log_df.loc[mask, "season"] = derived[mask]

    return df, log_df


def extract_workbook(path):
    wb = load_workbook(path, data_only=True)
    data_frames = []
    logs = []

    for ws in wb.worksheets:
        try:
            df_sheet, log = extract_sheet(ws, path)
            data_frames.append(df_sheet)
            logs.append(log)
        except Exception as e:
            logs.append(
                {
                    "source_file": Path(path).name,
                    "sheet_name": ws.title,
                    "season": season_from_path(path),
                    "error": repr(e),
                }
            )

    df = pd.concat(data_frames, ignore_index=True) if data_frames else pd.DataFrame()
    log_df = pd.DataFrame(logs)

    df, log_df = apply_season_date_fallback(df, log_df, path)

    return df, log_df


def extract_many(paths, fasta_path=None):
    all_data = []
    all_logs = []

    for path in paths:
        df, log_df = extract_workbook(path)
        all_data.append(df)
        all_logs.append(log_df)

    data = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()
    logs = pd.concat(all_logs, ignore_index=True) if all_logs else pd.DataFrame()

    if fasta_path is not None and not data.empty:
        seq_map = read_fasta_sequences(fasta_path)

        data["virus_seq"] = data["virus"].apply(lambda x: lookup_sequence(x, seq_map))
        data["reference_seq"] = data["reference"].apply(lambda x: lookup_sequence(x, seq_map))

        logs.attrs["fasta_log"] = {
            "fasta_path": str(fasta_path),
            "n_fasta_sequences": len(seq_map),
            "n_unique_virus_names": int(data["virus"].nunique(dropna=True)),
            "n_unique_reference_names": int(data["reference"].nunique(dropna=True)),
            "n_unique_virus_names_matched": int(data.loc[data["virus_seq"].notna(), "virus"].nunique(dropna=True)),
            "n_unique_reference_names_matched": int(data.loc[data["reference_seq"].notna(), "reference"].nunique(dropna=True)),
        }

    elif not data.empty:
        data["virus_seq"] = pd.NA
        data["reference_seq"] = pd.NA

    return data, logs


# -----------------------------
# Scoring and filtering
# -----------------------------

def is_nonempty_sequence(x) -> bool:
    """
    Return True when a sequence value is not missing and not an empty string.
    """
    if pd.isna(x):
        return False
    return bool(str(x).strip())


def count_x_in_sequence(x) -> int:
    """
    Count ambiguous amino acids represented by X/x in a sequence.
    Missing values are treated as having zero X, but missing sequences are
    filtered separately before this function is used for exclusion.
    """
    if pd.isna(x):
        return 0
    return str(x).upper().count("X")


def make_strain_passage_id(strain, passage_type) -> Optional[str]:
    """
    Build the strain+passage-type identifier used for strict self-titre matching.

    The score calculation follows the previous notebook logic:
        - compare within the same assay date
        - compare within the same reference strain and reference passage type
        - use self titre rows where virus_strain_passage == reference_strain_passage
    """
    strain_s = cell_to_str(strain)
    passage_s = cell_to_str(passage_type)

    if not strain_s or not passage_s:
        return None

    return f"{strain_s}_{passage_s}"


def build_score_filter_log(metrics: dict) -> pd.DataFrame:
    """
    Convert a metrics dictionary into a simple two-column QC log table.
    """
    return pd.DataFrame(
        [{"metric": key, "value": value} for key, value in metrics.items()]
    )


def build_scored_filtered_dataset(
    df: pd.DataFrame,
    fixed_score_scale: float = 8.0,
    titre_log_diff_violate_cutoff: float = -1.0,
    titre_log_diff_violate_rate_threshold: float = 0.5,
    remove_target_rows_linked_to_low_self_reference_groups: bool = True,
    low_self_titre_log_margin: float = 2.0,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build the filtered, scored H1N1 WHO HI dataset.

    Filters:
        1. Remove exact duplicate rows.
        2. Keep rows with usable numeric HI_titre values.
        3. Remove rows lacking virus_seq or reference_seq.
        4. Remove rows whose virus_seq or reference_seq contains >=4 X residues.
        5. Calculate titre_log_diff using strict self-titre controls.
        6. Remove rows for which a strict self-titre control is unavailable.
        7. For each date/reference/reference_passage_type group, calculate the
           fraction of rows with titre_log_diff <= titre_log_diff_violate_cutoff.
           If the fraction is greater than titre_log_diff_violate_rate_threshold,
           remove the entire group.
        8. If enabled, remove target-virus rows linked to bad reference groups
           with unusually low homologous/self titres.
        9. Remove rows with titre_log_diff < -1.

    Score definition:
        titre_log = log2(HI_titre / 10)
        titre_log_diff = mean(self titre_log in same date/reference_passage_type) - titre_log
        group-level and low-self target-side QC are applied before the row-level titre_log_diff filter
        rows with titre_log_diff < -1 are removed
        titre_log_diff_nonnegative = max(titre_log_diff, 0)
        score = min(titre_log_diff_nonnegative, fixed_score_scale) / fixed_score_scale

    The final score therefore remains in [0, 1] using a fixed 8-unit scale.
    """
    metrics = {}

    if df is None or df.empty:
        empty = pd.DataFrame()
        metrics["input_rows"] = 0
        metrics["final_rows"] = 0
        return empty, build_score_filter_log(metrics)

    data = df.copy()
    metrics["input_rows"] = int(len(data))

    # Keep raw passage histories before replacing final passage columns with passage types.
    data["virus_passage_raw"] = data.get("virus_passage", pd.NA)
    data["reference_passage_raw"] = data.get("reference_passage", pd.NA)

    metrics["unknown_virus_passage_type_rows_input"] = int(
        (data.get("virus_passage_type", pd.Series(index=data.index, dtype="object")) == "UNKNOWN").sum()
    )
    metrics["unknown_reference_passage_type_rows_input"] = int(
        (data.get("reference_passage_type", pd.Series(index=data.index, dtype="object")) == "UNKNOWN").sum()
    )
    metrics["unknown_either_passage_type_rows_input"] = int(
        (
            (data.get("virus_passage_type", pd.Series(index=data.index, dtype="object")) == "UNKNOWN")
            | (data.get("reference_passage_type", pd.Series(index=data.index, dtype="object")) == "UNKNOWN")
        ).sum()
    )

    before = len(data)
    data = data.drop_duplicates().reset_index(drop=True)
    metrics["dropped_exact_duplicate_rows"] = int(before - len(data))
    metrics["after_drop_exact_duplicates_rows"] = int(len(data))

    # HI_titre should already be parsed by parse_hi_value(), but enforce numeric values here.
    data["HI_titre"] = pd.to_numeric(data.get("HI_titre"), errors="coerce")
    before = len(data)
    data = data[data["HI_titre"].notna() & (data["HI_titre"] > 0)].copy()
    metrics["dropped_missing_or_invalid_HI_titre_rows"] = int(before - len(data))
    metrics["after_valid_HI_titre_rows"] = int(len(data))

    # Keep only records with both virus and reference sequences.
    virus_seq_present = data["virus_seq"].apply(is_nonempty_sequence)
    reference_seq_present = data["reference_seq"].apply(is_nonempty_sequence)
    metrics["rows_missing_virus_seq_before_sequence_filter"] = int((~virus_seq_present).sum())
    metrics["rows_missing_reference_seq_before_sequence_filter"] = int((~reference_seq_present).sum())
    metrics["rows_missing_either_seq_before_sequence_filter"] = int((~(virus_seq_present & reference_seq_present)).sum())

    before = len(data)
    data = data[virus_seq_present & reference_seq_present].copy()
    metrics["dropped_missing_sequence_rows"] = int(before - len(data))
    metrics["after_drop_missing_sequences_rows"] = int(len(data))

    # Count X residues and remove rows with >=4 X in either sequence.
    data["virus_seq_X_count"] = data["virus_seq"].apply(count_x_in_sequence)
    data["reference_seq_X_count"] = data["reference_seq"].apply(count_x_in_sequence)

    virus_x_bad = data["virus_seq_X_count"] >= 4
    reference_x_bad = data["reference_seq_X_count"] >= 4
    either_x_bad = virus_x_bad | reference_x_bad

    metrics["rows_with_virus_seq_X_count_ge_4"] = int(virus_x_bad.sum())
    metrics["rows_with_reference_seq_X_count_ge_4"] = int(reference_x_bad.sum())
    metrics["rows_with_either_seq_X_count_ge_4"] = int(either_x_bad.sum())

    before = len(data)
    data = data[~either_x_bad].copy()
    metrics["dropped_X_count_ge_4_rows"] = int(before - len(data))
    metrics["after_X_count_filter_rows"] = int(len(data))

    # Score calculation.
    data["titre_log"] = np.log2(data["HI_titre"] / 10)

    data["virus_strain_passage"] = [
        make_strain_passage_id(v, p)
        for v, p in zip(data["virus"], data["virus_passage_type"])
    ]
    data["reference_strain_passage"] = [
        make_strain_passage_id(v, p)
        for v, p in zip(data["reference"], data["reference_passage_type"])
    ]

    strict_self = data[data["virus_strain_passage"] == data["reference_strain_passage"]].copy()
    metrics["strict_self_titre_rows_after_filters"] = int(len(strict_self))
    metrics["strict_self_titre_groups_after_filters"] = int(
        strict_self[["date", "reference_strain_passage"]].drop_duplicates().shape[0]
    ) if not strict_self.empty else 0

    self_mean = (
        strict_self.groupby(["date", "reference_strain_passage"], dropna=False)["titre_log"]
        .mean()
        .reset_index()
        .rename(columns={"titre_log": "self_mean_titre_log"})
    )

    data = data.merge(
        self_mean,
        on=["date", "reference_strain_passage"],
        how="left",
    )

    data["titre_log_diff"] = data["self_mean_titre_log"] - data["titre_log"]

    before = len(data)
    data = data[data["titre_log_diff"].notna()].copy()
    metrics["dropped_rows_without_strict_self_titre"] = int(before - len(data))
    metrics["after_strict_self_titre_filter_rows"] = int(len(data))

    # Group-level QC: remove an entire date/reference/reference_passage group when
    # too many rows have titre_log_diff <= the violation cutoff. This catches
    # serum/date groups where many heterologous titres are higher than the self
    # titre, before the row-level titre_log_diff filter is applied.
    metrics["titre_log_diff_violate_cutoff"] = float(titre_log_diff_violate_cutoff)
    metrics["titre_log_diff_violate_rate_threshold"] = float(titre_log_diff_violate_rate_threshold)
    metrics["remove_target_rows_linked_to_low_self_reference_groups"] = bool(
        remove_target_rows_linked_to_low_self_reference_groups
    )
    metrics["low_self_titre_log_margin"] = float(low_self_titre_log_margin)

    if not 0 <= titre_log_diff_violate_rate_threshold <= 1:
        raise ValueError(
            "titre_log_diff_violate_rate_threshold must be between 0 and 1. "
            f"Got: {titre_log_diff_violate_rate_threshold}"
        )

    reference_passage_group_col = (
        "reference_passage_type" if "reference_passage_type" in data.columns else "reference_passage"
    )
    group_cols = ["date", "reference", reference_passage_group_col]
    metrics["titre_log_diff_violate_group_cols"] = ";".join(group_cols)

    if data.empty:
        data["titre_log_diff_group_n"] = pd.Series(dtype="int64")
        data["titre_log_diff_violate_n"] = pd.Series(dtype="int64")
        data["titre_log_diff_violate_rate"] = pd.Series(dtype="float64")
        data["titre_log_diff_viorate_rate"] = pd.Series(dtype="float64")
        data["titre_log_diff_group_removed_by_violate_rate"] = pd.Series(dtype="bool")
        data["date_median_self_mean_titre_log"] = pd.Series(dtype="float64")
        data["low_self_reference_group"] = pd.Series(dtype="bool")
        data["target_row_removed_by_low_self_reference_group"] = pd.Series(dtype="bool")
        metrics["titre_log_diff_violate_groups_before_filter"] = 0
        metrics["titre_log_diff_groups_removed_by_violate_rate"] = 0
        metrics["low_self_reference_groups_flagged_for_target_filter"] = 0
        metrics["rows_removed_by_titre_log_diff_violate_rate"] = 0
        metrics["rows_removed_by_low_self_reference_target_filter"] = 0
    else:
        data["titre_log_diff_is_violation"] = data["titre_log_diff"] <= titre_log_diff_violate_cutoff
        group_stats = (
            data.groupby(group_cols, dropna=False)
            .agg(
                titre_log_diff_group_n=("titre_log_diff", "size"),
                titre_log_diff_violate_n=("titre_log_diff_is_violation", "sum"),
                reference_group_self_mean_titre_log=("self_mean_titre_log", "first"),
            )
            .reset_index()
        )
        group_stats["titre_log_diff_violate_rate"] = (
            group_stats["titre_log_diff_violate_n"] / group_stats["titre_log_diff_group_n"]
        )
        group_stats["titre_log_diff_viorate_rate"] = group_stats["titre_log_diff_violate_rate"]
        group_stats["titre_log_diff_group_removed_by_violate_rate"] = (
            group_stats["titre_log_diff_violate_rate"] > titre_log_diff_violate_rate_threshold
        )

        metrics["titre_log_diff_violate_groups_before_filter"] = int(len(group_stats))
        metrics["titre_log_diff_groups_removed_by_violate_rate"] = int(
            group_stats["titre_log_diff_group_removed_by_violate_rate"].sum()
        )

        # Target-side QC for low-reactive homologous viruses:
        # If a bad reference group also has a self titre that is much lower than
        # the same-date median self titre, the problem is likely caused by the
        # virus/target strain itself rather than only by the serum column. In that
        # case, remove rows where the same strain appears as the target virus in
        # the same date and passage context. This is enabled by default.
        date_medians = (
            group_stats[["date", "reference_group_self_mean_titre_log"]]
            .dropna(subset=["reference_group_self_mean_titre_log"])
            .groupby("date", dropna=False)["reference_group_self_mean_titre_log"]
            .median()
            .reset_index()
            .rename(columns={"reference_group_self_mean_titre_log": "date_median_self_mean_titre_log"})
        )
        group_stats = group_stats.merge(date_medians, on="date", how="left")
        group_stats["low_self_reference_group"] = (
            group_stats["titre_log_diff_group_removed_by_violate_rate"]
            & group_stats["reference_group_self_mean_titre_log"].notna()
            & group_stats["date_median_self_mean_titre_log"].notna()
            & (
                group_stats["reference_group_self_mean_titre_log"]
                <= group_stats["date_median_self_mean_titre_log"] - low_self_titre_log_margin
            )
        )

        if not remove_target_rows_linked_to_low_self_reference_groups:
            group_stats["low_self_reference_group"] = False

        metrics["low_self_reference_groups_flagged_for_target_filter"] = int(
            group_stats["low_self_reference_group"].sum()
        )

        data = data.merge(group_stats, on=group_cols, how="left")

        virus_passage_group_col = (
            "virus_passage_type" if "virus_passage_type" in data.columns else "virus_passage"
        )
        low_self_targets = group_stats[group_stats["low_self_reference_group"]][
            ["date", "reference", reference_passage_group_col]
        ].rename(
            columns={
                "reference": "virus",
                reference_passage_group_col: virus_passage_group_col,
            }
        )
        low_self_targets = low_self_targets.drop_duplicates()
        low_self_targets["target_row_removed_by_low_self_reference_group"] = True

        data = data.merge(
            low_self_targets,
            on=["date", "virus", virus_passage_group_col],
            how="left",
        )
        data["target_row_removed_by_low_self_reference_group"] = data[
            "target_row_removed_by_low_self_reference_group"
        ].map(lambda x: bool(x) if pd.notna(x) else False)

        group_remove_mask = data["titre_log_diff_group_removed_by_violate_rate"].eq(True)
        target_remove_mask = data["target_row_removed_by_low_self_reference_group"].eq(True)
        before = len(data)
        metrics["rows_removed_by_titre_log_diff_violate_rate"] = int(group_remove_mask.sum())
        metrics["rows_removed_by_low_self_reference_target_filter"] = int(
            (target_remove_mask & ~group_remove_mask).sum()
        )
        data = data[~(group_remove_mask | target_remove_mask)].copy()
        metrics["after_titre_log_diff_violate_rate_filter_rows"] = int(len(data))
        metrics["after_low_self_reference_target_filter_rows"] = int(len(data))
        data = data.drop(columns=["titre_log_diff_is_violation"], errors="ignore")

    # Match the notebook filtering rule: remove rows where titre_log_diff < -1.
    before = len(data)
    data = data[data["titre_log_diff"] >= -1].copy()
    metrics["dropped_rows_titre_log_diff_lt_minus1"] = int(before - len(data))
    metrics["after_titre_log_diff_ge_minus1_filter_rows"] = int(len(data))

    data["titre_log_diff_nonnegative"] = data["titre_log_diff"].clip(lower=0)
    data["score"] = data["titre_log_diff_nonnegative"].clip(upper=fixed_score_scale) / fixed_score_scale

    metrics["fixed_score_scale"] = float(fixed_score_scale)
    metrics["max_titre_log_diff_nonnegative_final"] = float(data["titre_log_diff_nonnegative"].max()) if not data.empty else None
    metrics["n_rows_score_clipped_at_fixed_scale"] = int((data["titre_log_diff_nonnegative"] > fixed_score_scale).sum()) if not data.empty else 0
    metrics["max_score_final"] = float(data["score"].max()) if not data.empty else None

    metrics["unknown_virus_passage_type_rows_final"] = int((data["virus_passage_type"] == "UNKNOWN").sum())
    metrics["unknown_reference_passage_type_rows_final"] = int((data["reference_passage_type"] == "UNKNOWN").sum())
    metrics["unknown_either_passage_type_rows_final"] = int(
        ((data["virus_passage_type"] == "UNKNOWN") | (data["reference_passage_type"] == "UNKNOWN")).sum()
    )

    # Final requested output names: passage columns should contain passage type.
    data["virus_passage"] = data["virus_passage_type"]
    data["reference_passage"] = data["reference_passage_type"]

    requested_columns = [
        "season",
        "date",
        "virus",
        "reference",
        "virus_passage",
        "reference_passage",
        "virus_collection_date",
        "score",
        "censor",
        "virus_seq",
        "reference_seq",
    ]

    qc_columns = [
        "source_file",
        "sheet_name",
        "date_is_season_fallback",
        "virus_passage_raw",
        "reference_passage_raw",
        "virus_passage_type",
        "reference_passage_type",
        "HI_titre",
        "HI_raw",
        "titre_log",
        "self_mean_titre_log",
        "titre_log_diff",
        "titre_log_diff_group_n",
        "titre_log_diff_violate_n",
        "titre_log_diff_violate_rate",
        "titre_log_diff_viorate_rate",
        "titre_log_diff_group_removed_by_violate_rate",
        "reference_group_self_mean_titre_log",
        "date_median_self_mean_titre_log",
        "low_self_reference_group",
        "target_row_removed_by_low_self_reference_group",
        "titre_log_diff_nonnegative",
        "virus_seq_X_count",
        "reference_seq_X_count",
        "reference_from_header",
        "reference_from_row",
        "reference_match_score",
        "virus_collection_date_is_inferred_from_name",
        "row",
        "col",
        "virus_strain_passage",
        "reference_strain_passage",
    ]

    ordered = [c for c in requested_columns + qc_columns if c in data.columns]
    remaining = [c for c in data.columns if c not in ordered]
    data = data[ordered + remaining].reset_index(drop=True)

    metrics["final_rows"] = int(len(data))

    return data, build_score_filter_log(metrics)


# -----------------------------
# CLI
# -----------------------------

def resolve_excel_paths(excel_inputs: Iterable[str]) -> list[Path]:
    """
    Resolve --excel arguments into a sorted list of .xlsx files.
    """
    paths = []

    for item in excel_inputs:
        item = str(item)
        expanded = Path(item).expanduser()

        if any(ch in item for ch in "*?[]"):
            matches = [Path(x) for x in glob.glob(item)]
        elif expanded.is_dir():
            matches = list(expanded.glob("*.xlsx"))
        elif expanded.is_file():
            matches = [expanded]
        else:
            matches = [Path(x) for x in glob.glob(item)]

        for m in matches:
            if m.suffix.lower() == ".xlsx" and not m.name.startswith("~$"):
                paths.append(m)

    # Deduplicate while preserving sorted order.
    return sorted({str(p): p for p in paths}.values(), key=lambda x: str(x))


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Extract WHO HI assay Excel workbooks into a long-format CSV."
    )

    parser.add_argument(
        "--excel",
        nargs="+",
        default=["*.xlsx"],
        help=(
            "Excel file(s), directory/directories, or glob pattern(s). "
            "Examples: --excel WHO_data/NH2010.xlsx ; --excel WHO_data ; --excel 'WHO_data/*.xlsx'"
        ),
    )

    parser.add_argument(
        "--fasta",
        default=None,
        help=(
            "Optional FASTA file. Supports both H1N1-style and H3N2-style "
            "pipe-delimited headers; strain name is auto-detected."
        ),
    )

    parser.add_argument(
        "--out",
        default="WHO_HI_long_format.csv",
        help="Output CSV path for extracted long-format HI data.",
    )

    parser.add_argument(
        "--log",
        default="WHO_HI_sheet_log.csv",
        help="Output CSV path for per-sheet extraction log.",
    )

    parser.add_argument(
        "--score-out",
        default="H1N1_WHO_HI_long_format_with_score_filtered.csv",
        help="Output CSV path for filtered long-format HI data with fixed-scale score.",
    )

    parser.add_argument(
        "--score-log",
        default="H1N1_WHO_HI_score_filter_log.csv",
        help="Output CSV path for filtering/scoring QC metrics.",
    )

    parser.add_argument(
        "--score-scale",
        type=float,
        default=8.0,
        help="Fixed scale used to convert nonnegative titre-log difference to score. Default: 8.0",
    )

    parser.add_argument(
        "--titre-log-diff-violate-cutoff",
        type=float,
        default=-1.0,
        help=(
            "Cutoff used for group-level QC. Rows with titre_log_diff <= this value "
            "are counted as violations. Default: -1.0"
        ),
    )

    parser.add_argument(
        "--titre-log-diff-violate-rate-threshold",
        type=float,
        default=0.5,
        help=(
            "Remove all rows from a date/reference/reference_passage group when the "
            "fraction of rows with titre_log_diff <= --titre-log-diff-violate-cutoff "
            "is greater than this threshold. Default: 0.5"
        ),
    )
    parser.add_argument(
        "--remove-target-rows-linked-to-low-self-reference-groups",
        dest="remove_target_rows_linked_to_low_self_reference_groups",
        action="store_true",
        default=True,
        help=(
            "Also remove rows where a low-self bad reference strain appears as the "
            "target virus in the same date/passage context. This is enabled by default."
        ),
    )
    parser.add_argument(
        "--no-remove-target-rows-linked-to-low-self-reference-groups",
        dest="remove_target_rows_linked_to_low_self_reference_groups",
        action="store_false",
        help="Disable target-side removal linked to low-self bad reference groups.",
    )
    parser.add_argument(
        "--low-self-titre-log-margin",
        type=float,
        default=2.0,
        help=(
            "A bad reference group is considered low-self when its self_mean_titre_log "
            "is at least this many log2 units below the same-date median self_mean_titre_log. "
            "Default: 2.0"
        ),
    )

    parser.add_argument(
        "--drop-qc",
        action="store_true",
        help="Drop QC columns HI_raw, row, and col from the main output CSV.",
    )

    args = parser.parse_args(argv)

    input_files = resolve_excel_paths(args.excel)

    if not input_files:
        print(
            "No Excel workbook was found. Check the path passed to --excel.\n"
            f"Received --excel: {args.excel}",
            file=sys.stderr,
        )
        return 1

    fasta_path = Path(args.fasta).expanduser() if args.fasta else None

    if fasta_path is not None and not fasta_path.is_file():
        print(f"FASTA file was not found: {fasta_path}", file=sys.stderr)
        return 1

    df, sheet_log = extract_many(input_files, fasta_path=fasta_path)
    scored_df, score_log = build_scored_filtered_dataset(
        df,
        fixed_score_scale=args.score_scale,
        titre_log_diff_violate_cutoff=args.titre_log_diff_violate_cutoff,
        titre_log_diff_violate_rate_threshold=args.titre_log_diff_violate_rate_threshold,
        remove_target_rows_linked_to_low_self_reference_groups=(
            args.remove_target_rows_linked_to_low_self_reference_groups
        ),
        low_self_titre_log_margin=args.low_self_titre_log_margin,
    )

    main_df = df.copy()

    if args.drop_qc and not main_df.empty:
        main_df = main_df.drop(columns=["HI_raw", "row", "col"], errors="ignore")

    main_df.to_csv(args.out, index=False)
    sheet_log.to_csv(args.log, index=False)
    scored_df.to_csv(args.score_out, index=False)
    score_log.to_csv(args.score_log, index=False)

    n_error = 0

    if not sheet_log.empty and "error" in sheet_log.columns:
        n_error = int(sheet_log["error"].notna().sum())

    print(f"Extracted {len(df):,} rows from {len(input_files)} workbook(s).")
    print(f"Data written to: {args.out}")
    print(f"Log written to: {args.log}")
    print(f"Filtered/scored data written to: {args.score_out}")
    print(f"Filtering/scoring log written to: {args.score_log}")

    if n_error:
        print(
            f"Warning: {n_error} sheet(s) had extraction errors. Please check {args.log}.",
            file=sys.stderr,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
